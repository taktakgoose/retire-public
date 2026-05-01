#!/usr/bin/env python3
"""
retirement_to_xlsx.py - Convert retirement.csv to a formatted retirement.xlsx.

Usage:
    python3 retirement_to_xlsx.py [input.csv] [output.xlsx]

Defaults to retirement.csv -> retirement.xlsx in the current directory.

The CSV has two sections separated by a blank line:
  1. Year-by-year simulation data (header row + data rows)
  2. Summary block (label ; Ralph ; Sarah ; Total rows)
"""

import os
import re
import sys
import csv
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.chart import LineChart, BarChart, AreaChart, Reference, Series

# ---------------------------------------------------------------------------
# Colour palette
# ---------------------------------------------------------------------------
COLOUR_HEADER_BG   = "1F4E79"   # dark blue
COLOUR_HEADER_FG   = "FFFFFF"   # white
COLOUR_ALT_ROW     = "F5F0E8"   # light beige alternating row
COLOUR_SUMMARY_HDR = "2E75B6"   # medium blue for summary header
COLOUR_SUMMARY_BG  = "EBF3FB"   # very light blue for summary rows
COLOUR_NETWORTH_BG = "C6EFCE"   # green highlight for net-worth rows

FONT_NAME = "Arial"

# ---------------------------------------------------------------------------
# Strategy success thresholds (eval-pass "X of Y paths survived")
#
# The standard retirement-planning yardstick (Bengen / Trinity / FIRECalc) is
# a 95% success rate.  Below that the strategy starts trading safety margin
# for higher terminal wealth — still usable if you have backup flexibility
# (downsize, work longer, spending cuts), but the lower the number the more
# actively you need to manage sequence-of-returns risk.
# ---------------------------------------------------------------------------
SUCCESS_THRESHOLDS = [
    # (min_pct, label,      bg_colour, fg_colour, interpretation)
    (0.95,  "STRONG",      "107C10", "FFFFFF",
        "At or above the 95% planning benchmark — strategy survives even "
        "unfavourable market sequences."),
    (0.85,  "GOOD",        "5AA02C", "FFFFFF",
        "Above the 85% threshold — acceptable if you have meaningful "
        "flexibility (downsizing, part-time work, spending cuts)."),
    (0.70,  "MARGINAL",    "E8A33D", "000000",
        "Below the 85% planning threshold — workable but sensitive to bad "
        "market sequences.  Consider later moonbrook sale, lower adjustments, "
        "or a delayed retirement date."),
    (0.50,  "WEAK",        "E07020", "FFFFFF",
        "Under 70% survival — the strategy fails in a meaningful fraction "
        "of market paths.  Revisit assumptions before treating it as a plan."),
    (0.00,  "AT RISK",     "C00000", "FFFFFF",
        "Under 50% — the winning strategy loses more than half the time.  "
        "Not a viable retirement plan as modelled."),
]


def classify_success_rate(rate):
    """Return the (label, bg, fg, interpretation) tuple for a ratio in [0,1]."""
    for min_pct, label, bg, fg, interp in SUCCESS_THRESHOLDS:
        if rate >= min_pct:
            return label, bg, fg, interp
    return SUCCESS_THRESHOLDS[-1][1:]


def parse_success_ratio(summary_rows):
    """Extract (successes, total) from the 'Winning Strategy Successes' row.

    Returns (None, None) if the row is missing or malformed — the callout
    should then be skipped rather than showing garbage.
    """
    for row in summary_rows:
        if not row or row[0].strip() != "Winning Strategy Successes":
            continue
        # Value lives in column D (index 3) — format "<succ>/<total>".
        value = ""
        for cell in row[1:]:
            s = cell.strip()
            if s:
                value = s
                break
        m = re.match(r'^(\d+)\s*/\s*(\d+)$', value)
        if m:
            return int(m.group(1)), int(m.group(2))
    return None, None

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def make_font(bold=False, color="000000", size=10):
    return Font(name=FONT_NAME, bold=bold, color=color, size=size)

def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def try_numeric(value):
    """Return float/int if value looks numeric, otherwise the raw string."""
    v = value.strip()
    if v == "" or v == "-":
        return v
    try:
        f = float(v)
        return int(f) if f.is_integer() else f
    except ValueError:
        return v

def split_sections(rows):
    """Split raw CSV rows into (data_rows, summary_rows) at the first blank row."""
    split_at = None
    for idx, row in enumerate(rows):
        if all(cell.strip() == "" for cell in row):
            split_at = idx
            break
    if split_at is None:
        return rows, []
    return rows[:split_at], rows[split_at + 1:]

# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# Section layout for the Simulation sheet
# ---------------------------------------------------------------------------

# Each entry maps a header label that *starts* a section to a (display name,
# background colour) pair.  "Asset" is handled separately (first occurrence).
SECTION_STARTS = [
    ("Year",         ("Year Overview",            "1F4E79")),
    ("Ralph RRSP",   ("Ralph \u2014 Withdrawals", "1E5631")),
    ("Sarah RRSP",    ("Sarah \u2014 Withdrawals",  "154360")),
    ("Gross",        ("Net Worth",                "4A235A")),
    ("Ralph Age",    ("Ralph \u2014 Tax Detail",  "1E5631")),
    # "Sarah Age" appears twice: once in the year-overview block and again as
    # the first column of the Sarah tax-detail block.  identify_sections uses
    # the LAST occurrence of any duplicated label, so this correctly lands on
    # the tax-detail column rather than the earlier overview column.
    ("Sarah Age",     ("Sarah \u2014 Tax Detail",   "154360")),
    # "Asset" sentinel handled below
]

def identify_sections(header):
    """Return [(label, colour, start_col_1based, end_col_1based), ...]."""
    stripped = [h.strip() for h in header]
    lookup   = {label: info for label, info in SECTION_STARTS}

    # Build a map from section label → last 0-based column index.
    # Using the last occurrence means that labels shared with an earlier
    # overview column (e.g. "Sarah Age") still resolve to the correct section.
    last_pos = {}
    for i, h in enumerate(stripped):
        if h in lookup:
            last_pos[h] = i

    boundaries = []   # (col_0based, label, colour)
    for label, (display, colour) in lookup.items():
        if label in last_pos:
            boundaries.append((last_pos[label], display, colour))

    # "Asset" sentinel: first occurrence not already claimed as a section start.
    for i, h in enumerate(stripped):
        if h == "Asset" and not any(b[1] == "Assets" for b in boundaries):
            boundaries.append((i, "Assets", "614B0A"))
            break

    boundaries.sort(key=lambda x: x[0])

    sections = []
    for i, (col0, label, colour) in enumerate(boundaries):
        start = col0 + 1                                        # 1-based
        end   = boundaries[i + 1][0] if i + 1 < len(boundaries) else len(stripped)
        sections.append((label, colour, start, end))
    return sections


def build_data_sheet(ws, data_rows):
    if not data_rows:
        return

    header = data_rows[0]
    body   = data_rows[1:]

    sections = identify_sections(header)

    # Columns that start a new section — receive a thick left border.
    section_start_cols = {s[2] for s in sections}

    # Columns formatted as percentages.  The C simulation writes rates as
    # fractions (e.g. 0.0683); without the '%' format Excel rounds them to 0
    # under the default '#,##0' format.  Split into two groups so we can use
    # tighter precision for the per-asset Growth columns (where Monte-Carlo
    # shocks produce variations down at the hundredth of a percent).
    tax_rate_cols = {
        col_idx
        for col_idx, label in enumerate(header, start=1)
        if "tax rate" in label.strip().lower()
    }
    growth_cols = {
        col_idx
        for col_idx, label in enumerate(header, start=1)
        if label.strip().lower() == "growth"
    }

    thick_left = Border(left=Side(style="medium", color="000000"))

    def apply_section_border(cell, col_idx):
        if col_idx in section_start_cols:
            cell.border = thick_left

    # --- Row 1: section banner ---
    for label, colour, start_col, end_col in sections:
        # Write label in the first cell of the section.
        cell = ws.cell(row=1, column=start_col, value=label)
        cell.font      = make_font(bold=True, color=COLOUR_HEADER_FG, size=10)
        cell.fill      = make_fill(colour)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=False)
        apply_section_border(cell, start_col)

        # Merge across the section width (if more than one column).
        if end_col > start_col:
            ws.merge_cells(start_row=1, start_column=start_col,
                           end_row=1,   end_column=end_col)

        # Fill remaining cells in the merged range so the colour is contiguous
        # even before merge is rendered (also ensures border carries through).
        for col_idx in range(start_col + 1, end_col + 1):
            c = ws.cell(row=1, column=col_idx)
            c.fill = make_fill(colour)

    ws.row_dimensions[1].height = 18

    # --- Row 2: column header ---
    for col_idx, label in enumerate(header, start=1):
        cell = ws.cell(row=2, column=col_idx, value=label.strip())
        cell.font      = make_font(bold=True, color=COLOUR_HEADER_FG, size=10)
        cell.fill      = make_fill(COLOUR_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        apply_section_border(cell, col_idx)

    # --- Rows 3+: data ---
    for row_idx, row in enumerate(body, start=3):
        fill = make_fill(COLOUR_ALT_ROW) if row_idx % 2 == 1 else None
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=try_numeric(value))
            cell.font = make_font(size=10)
            if fill:
                cell.fill = fill
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal="right")
                if col_idx in tax_rate_cols:
                    cell.number_format = '0.0%'
                elif col_idx in growth_cols:
                    cell.number_format = '0.00%'
                elif col_idx > 1:
                    cell.number_format = '#,##0'
            else:
                cell.alignment = Alignment(horizontal="left")
            apply_section_border(cell, col_idx)

    # --- Column widths ---
    ws.column_dimensions[get_column_letter(1)].width = 6   # Year
    for col_idx in range(2, len(header) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 11

    # --- Data bar conditional formatting (withdrawal / income columns) ---
    header_lower   = [h.strip().lower() for h in header]
    exclude_labels = {"robertson", "moonbrook"}
    num_data_rows  = len(body)
    try:
        DATABAR_FIRST_COL = header_lower.index("expenses") + 2
        DATABAR_LAST_COL  = header_lower.index("gross")
    except ValueError:
        DATABAR_FIRST_COL = 6
        DATABAR_LAST_COL  = 25

    included_cols = [
        get_column_letter(col_idx)
        for col_idx, label in enumerate(header, start=1)
        if DATABAR_FIRST_COL <= col_idx <= DATABAR_LAST_COL
        and label.strip().lower() not in exclude_labels
    ]

    if included_cols:
        multi_range = " ".join(
            f"{col}3:{col}{num_data_rows + 2}" for col in included_cols
        )
        ws.conditional_formatting.add(multi_range, DataBarRule(
            start_type="num", start_value=0,
            end_type="max",
            color="2E75B6",
        ))

    # --- Colour scale on Expenses column ---
    for col_idx, label in enumerate(header, start=1):
        if label.strip().lower() == "expenses":
            col_letter = get_column_letter(col_idx)
            data_range = f"{col_letter}3:{col_letter}{num_data_rows + 2}"
            ws.conditional_formatting.add(data_range, ColorScaleRule(
                start_type="min",        start_color="FF0000",
                mid_type="percentile",   mid_value=50, mid_color="FFFF00",
                end_type="max",          end_color="00B050",
            ))

    # Freeze section banner + column header rows, and the Year column.
    ws.freeze_panes = "B3"


def build_success_callout(ws, successes, total, start_row=1):
    """Render a 4-row prominent banner showing the eval-pass success ratio.

    Layout (each row spans columns A:D so the banner is visually distinct
    from the summary table that follows):
        Row N   — title "Strategy Robustness"  (callout-coloured background)
        Row N+1 — large "X / Y paths survived  (NN%)" text
        Row N+2 — threshold label ("STRONG", "MARGINAL", etc.)
        Row N+3 — interpretation paragraph (smaller italic)
    Returns the next free row number after the callout (caller should add
    one more blank spacer before the summary header).
    """
    rate = successes / total if total > 0 else 0.0
    label, bg, fg, interp = classify_success_rate(rate)
    pct_text = f"{rate*100:.0f}%"

    # --- Row N: title ---
    title = ws.cell(row=start_row, column=1, value="Strategy Robustness")
    title.font      = make_font(bold=True, color=fg, size=12)
    title.fill      = make_fill(bg)
    title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for c in range(2, 5):
        ws.cell(row=start_row, column=c).fill = make_fill(bg)
    ws.merge_cells(start_row=start_row, start_column=1,
                   end_row=start_row,   end_column=4)
    ws.row_dimensions[start_row].height = 22

    # --- Row N+1: big number ---
    big = ws.cell(row=start_row + 1, column=1,
                  value=f"{successes} / {total} market paths survived  ({pct_text})")
    big.font      = make_font(bold=True, color=fg, size=18)
    big.fill      = make_fill(bg)
    big.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for c in range(2, 5):
        ws.cell(row=start_row + 1, column=c).fill = make_fill(bg)
    ws.merge_cells(start_row=start_row + 1, start_column=1,
                   end_row=start_row + 1,   end_column=4)
    ws.row_dimensions[start_row + 1].height = 32

    # --- Row N+2: threshold label ---
    lbl = ws.cell(row=start_row + 2, column=1, value=label)
    lbl.font      = make_font(bold=True, color=fg, size=11)
    lbl.fill      = make_fill(bg)
    lbl.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for c in range(2, 5):
        ws.cell(row=start_row + 2, column=c).fill = make_fill(bg)
    ws.merge_cells(start_row=start_row + 2, start_column=1,
                   end_row=start_row + 2,   end_column=4)
    ws.row_dimensions[start_row + 2].height = 18

    # --- Row N+3: interpretation ---
    note = ws.cell(row=start_row + 3, column=1, value=interp)
    note.font      = Font(name=FONT_NAME, italic=True, color=fg, size=10)
    note.fill      = make_fill(bg)
    note.alignment = Alignment(horizontal="left", vertical="center",
                               indent=1, wrap_text=True)
    for c in range(2, 5):
        ws.cell(row=start_row + 3, column=c).fill = make_fill(bg)
    ws.merge_cells(start_row=start_row + 3, start_column=1,
                   end_row=start_row + 3,   end_column=4)
    ws.row_dimensions[start_row + 3].height = 36

    return start_row + 4


def build_summary_sheet(ws, summary_rows):
    if not summary_rows:
        return

    # First non-empty row is the column-label row (;Ralph;Sarah;Total;)
    label_row = None
    data_start = 0
    for idx, row in enumerate(summary_rows):
        if any(c.strip() for c in row):
            label_row  = row
            data_start = idx + 1
            break

    if label_row is None:
        return

    # --- Strategy-robustness callout (only when the eval-pass ratio is in the
    # CSV; older runs without the row simply skip the banner). ---
    successes, total = parse_success_ratio(summary_rows)
    next_row = 1
    if successes is not None and total:
        next_row = build_success_callout(ws, successes, total, start_row=1)
        next_row += 1   # blank spacer between callout and summary header
    header_row = next_row
    data_offset = header_row + 1   # data rows start one below the header

    # --- Header ---
    headers = [c.strip() for c in label_row]
    for col_idx, label in enumerate(headers, start=1):
        if not label:
            continue
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.font      = make_font(bold=True, color=COLOUR_HEADER_FG, size=10)
        cell.fill      = make_fill(COLOUR_SUMMARY_HDR)
        cell.alignment = Alignment(horizontal="center")

    # --- Data rows ---
    for row_idx, row in enumerate(summary_rows[data_start:], start=data_offset):
        if not any(c.strip() for c in row):
            continue
        label       = row[0] if row else ""
        is_networth = "Networth" in label
        # "Average …" rows emit tax-rate percentages.  The simulator also
        # emits "Expected/Realized <Class> Return" rows (as of 2026-04-17) —
        # those are fractional rates too, so they get the same treatment.
        is_tax_pct  = "Average"  in label
        is_ret_pct  = (("Expected" in label or "Realized" in label or
                        "Min" in label or "Max" in label)
                       and "Return" in label)
        is_percent  = is_tax_pct or is_ret_pct
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=try_numeric(value))
            cell.font = make_font(bold=is_networth, size=10)
            cell.fill = make_fill(COLOUR_NETWORTH_BG) if is_networth \
                        else make_fill(COLOUR_SUMMARY_BG)
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal="right")
                if is_tax_pct:
                    cell.number_format = '0.0%'
                elif is_ret_pct:
                    cell.number_format = '0.00%'
                else:
                    cell.number_format = '#,##0'
            else:
                cell.alignment = Alignment(horizontal="left")

    # --- Column widths ---
    ws.column_dimensions["A"].width = 26
    for col_letter in ["B", "C", "D"]:
        ws.column_dimensions[col_letter].width = 18

# ---------------------------------------------------------------------------
# Charts sheet
# ---------------------------------------------------------------------------

def build_charts_sheet(ws, data_rows, wb, bands=None):
    """
    Build three charts on a dedicated sheet, all referencing the Simulation
    sheet directly so they update if the source data changes.

    Chart 1 — Net Worth & Expenses over time  (line chart)
    Chart 2 — Annual Income Sources           (stacked bar chart)
    Chart 3 — Asset Values over time          (stacked area chart)

    A hidden helper sheet ("ChartData") holds the summed Ralph+Sarah asset
    values used by Chart 3 — Excel plots from hidden sheets fine, unlike
    hidden columns.
    """
    if not data_rows or len(data_rows) < 2:
        return

    SIM = "Simulation"
    header    = data_rows[0]
    num_rows  = len(data_rows) - 1   # exclude header
    # Simulation sheet layout: row 1 = section banners, row 2 = column headers,
    # rows 3..data_end = data.
    HDR_ROW  = 2
    DATA_ROW = 3
    data_end = num_rows + 2          # last data row (1-indexed)

    # Build a label->col_index lookup (1-based)
    col = {h.strip(): i for i, h in enumerate(header, start=1)}

    # -----------------------------------------------------------------------
    # Chart 1: Net Worth & Expenses over time (line chart)
    # -----------------------------------------------------------------------
    ch1 = LineChart()
    ch1.title       = "Net Worth & Expenses Over Time"
    ch1.style       = 10
    ch1.y_axis.title = "Amount ($)"
    ch1.x_axis.title = "Year"
    ch1.width        = 24
    ch1.height       = 14

    year_ref = Reference(wb[SIM],
                         min_col=col["Year"], min_row=DATA_ROW, max_row=data_end)

    for label, col_name in [
        ("Gross",      "Gross"),
        ("After-Tax",  "After-Tax"),
        ("Cash",       "Cash"),
        ("Expenses",   "Expenses"),
    ]:
        c = col[col_name]
        ref = Reference(wb[SIM], min_col=c, min_row=HDR_ROW, max_row=data_end)
        series = Series(ref, title_from_data=True)
        ch1.append(series)

    ch1.set_categories(year_ref)
    ch1.shape = 4
    ws.add_chart(ch1, "B2")

    # -----------------------------------------------------------------------
    # Chart 2: Annual Income Sources — stacked bar
    # Withdrawal debits for RRSP, DCPP, TFSA, NonReg, + income streams
    # (CPP, OAS, Rent, Polaron) drawn from cols F-Y.
    # -----------------------------------------------------------------------
    ch2 = BarChart()
    ch2.type         = "col"
    ch2.grouping     = "stacked"
    ch2.overlap      = 100
    ch2.title        = "Annual Income Sources"
    ch2.y_axis.title = "Amount ($)"
    ch2.x_axis.title = "Year"
    ch2.width        = 24
    ch2.height       = 14

    # Columns to include: F-Y (indices 6-25), grouped by friendly label.
    # Each entry: (display_label, header_name_in_csv)
    income_sources = [
        ("Ralph RRSP",   "Ralph RRSP"),
        ("Sarah RRSP",    "Sarah RRSP"),
        ("Ralph DCPP",   "DCPP"),          # first DCPP col = Ralph
        ("Sarah DCPP",    "DCPP"),          # second DCPP col = Sarah
        ("Ralph NonReg", "NonReg"),        # first NonReg col = Ralph
        ("Sarah NonReg",  "NonReg"),        # second NonReg col = Sarah
        ("Ralph TFSA",   "TFSA"),          # first TFSA col = Ralph
        ("Sarah TFSA",    "TFSA"),          # second TFSA col = Sarah
        ("Ralph CPP",    "CPP"),
        ("Sarah CPP",     "CPP"),
        ("Ralph OAS",    "OAS"),
        ("Sarah OAS",     "OAS"),
        ("Ralph Rent",   "Rent"),
        ("Sarah Rent",    "Rent"),
    ]

    # For duplicate header names (DCPP, NonReg etc.) track which occurrence
    # we've already used so we pick Ralph first then Sarah.
    seen_counts = {}
    header_occurrences = {}
    for i, h in enumerate(header, start=1):
        k = h.strip()
        header_occurrences.setdefault(k, []).append(i)

    # Write Chart 2 series data into ChartData so labels come from cell
    # values (title_from_data=True), which Excel renders reliably.
    # SeriesLabel(v=...) is unreliable across Excel versions.
    wh = wb.create_sheet("ChartData")
    wh.sheet_state = "hidden"

    # Col 1 of ChartData = Year
    wh.cell(row=1, column=1, value="Year")
    for r, data_row in enumerate(data_rows[1:], start=2):
        wh.cell(row=r, column=1,
                value=try_numeric(data_row[col["Year"] - 1]))

    # Re-walk income_sources, writing each series into ChartData and
    # adding it to ch2 with title_from_data=True.
    ch2_col = 1   # last ChartData column written for Chart 2
    seen_counts2 = {}
    for label, col_name in income_sources:
        seen_counts2[col_name] = seen_counts2.get(col_name, 0)
        occurrences = header_occurrences.get(col_name, [])
        valid = [c for c in occurrences if 6 <= c <= 25]
        if seen_counts2[col_name] >= len(valid):
            continue
        sim_col = valid[seen_counts2[col_name]]
        seen_counts2[col_name] += 1

        ch2_col += 1
        wh.cell(row=1, column=ch2_col, value=label)
        for r, data_row in enumerate(data_rows[1:], start=2):
            wh.cell(row=r, column=ch2_col,
                    value=try_numeric(data_row[sim_col - 1]) or 0)

        ref = Reference(wh, min_col=ch2_col, min_row=1, max_row=num_rows + 1)
        ch2.append(Series(ref, title_from_data=True))

    ch2.set_categories(Reference(wh, min_col=1, min_row=2, max_row=num_rows + 1))
    ws.add_chart(ch2, "B32")

    # -----------------------------------------------------------------------
    # Chart 3: Asset Values over time - stacked area
    # Continues writing into ChartData starting after the Chart 2 columns.
    # -----------------------------------------------------------------------
    ch3_col_offset = ch2_col   # Chart 3 columns start after Chart 2 data

    ch3 = AreaChart()
    ch3.grouping     = "stacked"
    ch3.title        = "Asset Values Over Time"
    ch3.y_axis.title = "Value ($)"
    ch3.x_axis.title = "Year"
    ch3.width        = 24
    ch3.height       = 14

    asset_groups = [
        ("RRSP",        ["RRSP"]),
        ("DCPP / LIRA", ["LIRA and DCPP"]),
        ("TFSA",        ["TFSA"]),
        ("Non-Reg",     ["NonReg"]),
        ("Properties",  ["Robertson", "Moonbrook Street"]),
    ]

    # Asset blocks: columns whose header is exactly "Asset" hold the asset name;
    # the immediately following column holds its value.  Scan the header rather
    # than relying on a hardcoded stride so column layout changes don't break this.
    asset_blocks = [
        (i + 1, i + 2)
        for i, h in enumerate(header)
        if h.strip() == "Asset"
    ]

    # Build asset_name -> [value_col, ...] from first data row
    first_data = data_rows[1]
    asset_name_to_value_cols = {}
    for name_col, value_col in asset_blocks:
        name = first_data[name_col - 1].strip() if name_col - 1 < len(first_data) else ""
        if name:
            asset_name_to_value_cols.setdefault(name, []).append(value_col)

    for helper_col, (group_label, keywords) in enumerate(asset_groups, start=ch3_col_offset + 1):
        matched_cols = [
            vc
            for asset_name, vcols in asset_name_to_value_cols.items()
            if any(kw in asset_name for kw in keywords)
            for vc in vcols
        ]
        if not matched_cols:
            continue

        wh.cell(row=1, column=helper_col, value=group_label)
        for r, data_row in enumerate(data_rows[1:], start=2):
            total = sum(
                (try_numeric(data_row[vc - 1]) or 0)
                for vc in matched_cols
                if vc - 1 < len(data_row)
            )
            wh.cell(row=r, column=helper_col, value=total)

        series = Series(
            Reference(wh, min_col=helper_col, min_row=1, max_row=num_rows + 1),
            title_from_data=True,
        )
        ch3.append(series)

    ch3.set_categories(Reference(wh, min_col=1, min_row=2, max_row=num_rows + 1))

    # Add Cash as the final series
    cash_helper_col = ch3_col_offset + len(asset_groups) + 1
    wh.cell(row=1, column=cash_helper_col, value="Cash")
    cash_col = col["Cash"]
    for r, data_row in enumerate(data_rows[1:], start=2):
        wh.cell(row=r, column=cash_helper_col,
                value=try_numeric(data_row[cash_col - 1]))
    series = Series(
        Reference(wh, min_col=cash_helper_col, min_row=1, max_row=num_rows + 1),
        title_from_data=True,
    )
    ch3.append(series)

    ws.add_chart(ch3, "B62")

    # -----------------------------------------------------------------------
    # Percentile confidence bands (P10 / P50 / P90) appended to ch1.
    # Values are written into ChartData columns after the asset/cash block
    # so Excel can reference them with a proper data range.
    # -----------------------------------------------------------------------
    if bands:
        band_specs = [
            ("P10 Net Worth", "NW_P10", "AAAAAA"),
            ("P50 Net Worth", "NW_P50", "555555"),
            ("P90 Net Worth", "NW_P90", "AAAAAA"),
        ]
        band_col_start = cash_helper_col + 1
        for b_offset, (b_label, b_key, b_colour) in enumerate(band_specs):
            b_vals = bands.get(b_key, [])
            if not b_vals:
                continue
            b_col = band_col_start + b_offset
            wh.cell(row=1, column=b_col, value=b_label)
            for r, v in enumerate(b_vals[:num_rows], start=2):
                wh.cell(row=r, column=b_col, value=v)
            ref = Reference(wh, min_col=b_col, min_row=1, max_row=num_rows + 1)
            band_series = Series(ref, title_from_data=True)
            band_series.graphicalProperties.line.solidFill = b_colour
            ch1.append(band_series)



# ---------------------------------------------------------------------------
# Assumptions sheet
# ---------------------------------------------------------------------------

# Sections shown on the Assumptions sheet.
# Each parameter entry: (params.h define name, display label, format type)
# format types: "dollar", "percent", "year"
ASSUMPTION_SECTIONS = [
    ("Simulation Period", [
        ("CURRENT_YEAR",            "Current Year",              "year"),
        ("CURRENT_MONTH",          "Simulation Start Month",    "count"),
        ("RALPH_RETIREMENT_YEAR",  "Ralph Retirement Year",     "year"),
        ("RALPH_RETIREMENT_MONTH", "Ralph Retirement Month",    "count"),
        ("SARAH_RETIREMENT_YEAR",   "Sarah Retirement Year",      "year"),
        ("SARAH_RETIREMENT_MONTH",  "Sarah Retirement Month",     "count"),
        ("PARTIAL_YEAR_MONTHS",    "Calendar Months in Start Year (expenses & growth scale)", "count"),
        ("DEATH_YEAR",             "Projected End Year",       "year"),
        ("N_TRIALS",               "Monte Carlo Trials",       "count"),
    ]),
    ("Income", [
        ("RALPH_SALARY",                    "Ralph Salary",               "dollar"),
        ("SARAH_SALARY",                     "Sarah Salary",                "dollar"),
        ("DCPP_EMPLOYER_CONTRIBUTION_RATE", "Employer DCPP Contribution", "percent"),
        ("RALPH_CPP_MONTHLY_PAYMENT",       "Ralph CPP Monthly Payment",  "dollar"),
        ("RALPH_CPP_START_AGE",             "Ralph CPP Start Age",        "age"),
        ("SARAH_CPP_MONTHLY_PAYMENT",        "Sarah CPP Monthly Payment",   "dollar"),
        ("SARAH_CPP_START_AGE",              "Sarah CPP Start Age",         "age"),
        ("RALPH_WORK_MONTHS_YEAR1", "Ralph Work Months (Year 1 salary proration)", "count"),
        ("SARAH_WORK_MONTHS_YEAR1",  "Sarah Work Months (Year 1 salary proration)",  "count"),
    ]),
    ("Starting Balance Totals", [
        ("STARTING_CASH",         "Cash",              "dollar"),
        ("STARTING_RALPH_RRSP",   "Ralph RRSP",        "dollar"),
        ("STARTING_SARAH_RRSP",    "Sarah RRSP",         "dollar"),
        ("STARTING_RALPH_DCPP",   "Ralph DCPP / LIRA", "dollar"),
        ("STARTING_SARAH_DCPP",    "Sarah DCPP / LIRA",  "dollar"),
        ("STARTING_RALPH_TFSA",   "Ralph TFSA",        "dollar"),
        ("STARTING_SARAH_TFSA",    "Sarah TFSA",         "dollar"),
        ("STARTING_RALPH_NONREG", "Ralph Non-Reg",     "dollar"),
    ]),
    ("Contribution Room", [
        ("STARTING_RALPH_RRSP_ROOM", "Ralph RRSP Room", "dollar"),
        ("STARTING_SARAH_RRSP_ROOM",  "Sarah RRSP Room",  "dollar"),
        ("STARTING_RALPH_TFSA_ROOM", "Ralph TFSA Room",  "dollar"),
        ("STARTING_SARAH_TFSA_ROOM",  "Sarah TFSA Room",   "dollar"),
    ]),
    ("Account Breakdown \u2014 Ralph", [
        ("GENERICBANK_RRSP_RALPH",           "GENERICBANK RRSP",            "dollar"),
        ("GENERICBROKER_RRSP_RALPH",   "Genericbroker RRSP",    "dollar"),
        ("GENERICBANK_LIRA_RALPH",           "GENERICBANK LIRA",            "dollar"),
        ("SUNLIFE_DCPP_RALPH",        "Sun Life DCPP",        "dollar"),
        ("GENERICBANK_TFSA_RALPH",           "GENERICBANK TFSA",            "dollar"),
        ("GENERICBROKER_TFSA_RALPH",   "Genericbroker TFSA",    "dollar"),
        ("GENERICBROKER_NONREG_RALPH", "Genericbroker Non-Reg", "dollar"),
        ("GENERICBROKER_CRYPTO_RALPH", "Genericbroker Crypto",  "dollar"),
    ]),
    ("Account Breakdown \u2014 Sarah", [
        ("GENERICBANK_RRSP_SARAH",            "GENERICBANK RRSP",            "dollar"),
        ("SUNLIFE_RRSP_SARAH",         "Sun Life RRSP",        "dollar"),
        ("GENERICBANK_LIRA_SARAH",            "GENERICBANK LIRA",            "dollar"),
        ("SUNLIFE_DCPP_SARAH",         "Sun Life DCPP",        "dollar"),
        ("GENERICBANK_TFSA_SARAH",            "GENERICBANK TFSA",            "dollar"),
        ("GENERICBROKER_TFSA_SARAH",    "Genericbroker TFSA",    "dollar"),
        ("WORKPLACE_BONUS_SARAH",          "Workplace Bonus",          "dollar"),
    ]),
    ("Cash & Liabilities", [
        ("GENERICBANK_CHEQUING_ACCT",    "GENERICBANK Chequing",             "dollar"),
        ("WORKPLACE_AFTER_TAX_STOCK", "Workplace After-Tax Stock",     "dollar"),
        ("CRA_REFUND",            "CRA Refund / (Owing)",      "dollar"),
        ("AMEX_BALANCE",          "Amex (liability)",          "dollar"),
        ("VISA_BALANCE",        "Visa LOC (liability)",    "dollar"),
        ("MASTERCARD_BALANCE",        "Mastercard (liability)",  "dollar"),
        ("DISCOVERY_BALANCE",        "Discovery Card (liability)",   "dollar"),
    ]),
    ("Growth & Inflation", [
        ("INFLATION",         "Inflation Rate",         "percent"),
        ("RALPH_GROWTH_RATE", "Ralph Portfolio Growth", "percent"),
        ("SARAH_GROWTH_RATE",  "Sarah Portfolio Growth",  "percent"),
    ]),
    # Expected (base) rates above are the drift applied every year.  Actual
    # year-to-year returns are drawn from a normal distribution around that
    # drift using the volatilities below — one independent draw per asset
    # class per year per simulated path.  RETURN_PATHS_PER_STRATEGY controls
    # how many distinct market paths each candidate withdrawal strategy is
    # evaluated against (higher = less luck-driven overfitting).
    ("Monte Carlo Volatility", [
        ("FINANCIAL_RETURN_VOLATILITY", "Financial Return Volatility (\u03c3)", "percent"),
        ("PROPERTY_RETURN_VOLATILITY",  "Property Return Volatility (\u03c3)",  "percent"),
        ("RENT_RETURN_VOLATILITY",      "Rent Return Volatility (\u03c3)",      "percent"),
        ("RETURN_PATHS_PER_STRATEGY",   "Return Paths per Strategy",            "count"),
    ]),
    ("Real Estate", [
        ("MOONBROOK_SALE_YEAR",               "Moonbrook Street Sale Year",          "year"),
        ("MOONBROOK_CURRENT_VALUE",           "Moonbrook Street Current Value",      "dollar"),
        ("MOONBROOK_PURCHASE_COST",           "Moonbrook Street Purchase Cost (ACB)","dollar"),
        ("MOONBROOK_UCC",                     "Moonbrook Street UCC",                "dollar"),
        ("MOONBROOK_ANNUAL_GROWTH",           "Moonbrook Street Annual Growth",      "percent"),
        ("MOONBROOK_ANNUAL_RENT_PROFIT",      "Moonbrook Annual Rent Profit",        "dollar"),
        ("MOONBROOK_MORTGAGE_BALANCE",        "Moonbrook Mortgage Balance",          "dollar"),
        ("MOONBROOK_MORTGAGE_MONTHLY_PAYMENT","Moonbrook Monthly Mortgage Payment",  "dollar"),
        ("MOONBROOK_MORTGAGE_INTEREST_RATE",  "Moonbrook Mortgage Interest Rate",    "percent"),
        ("ROBERTSON_SALE_YEAR",             "Robertson Court Sale Year",         "year"),
        ("ROBERTSON_CURRENT_VALUE",         "Robertson Court Current Value",     "dollar"),
        ("ROBERTSON_ANNUAL_GROWTH",         "Robertson Court Annual Growth",     "percent"),
        ("ROBERTSON_MORTGAGE_BALANCE",      "Robertson Mortgage Balance",        "dollar"),
        ("ROBERTSON_MORTGAGE_MONTHLY_PAYMENT","Robertson Monthly Mortgage Payment","dollar"),
        ("ROBERTSON_MORTGAGE_INTEREST_RATE","Robertson Mortgage Interest Rate",  "percent"),
    ]),
    ("Spending Phases", [
        ("KIDS_HOME_SPENDING",       "Current (Kids at Home)",        "dollar"),
        ("GO_GO_YEAR",               "Go-Go Phase Starts",            "year"),
        ("GO_GO_SPENDING",           "Go-Go Annual Spending",         "dollar"),
        ("SLOW_GO_YEAR",             "Slow-Go Phase Starts",          "year"),
        ("SLOW_GO_SPENDING",         "Slow-Go Annual Spending",       "dollar"),
        ("NO_GO_YEAR",               "No-Go Phase Starts",            "year"),
        ("NO_GO_SPENDING",           "No-Go Annual Spending",         "dollar"),
        ("ROBERTSON_SALE_YEAR",        "Senior Living Starts (After)",  "year"),
        ("SENIOR_LIVING_MONTHLY_RENT","Senior Living Monthly Rent",   "dollar"),
    ]),
    ("Withdrawal & Tax", [
        ("PREFERRED_CASH",              "Minimum Cash Balance",           "dollar"),
        ("NONREG_CASH_THRESHOLD",       "Non-Reg Investment Threshold",   "dollar"),
        ("LOWEST_TAX_THRESHOLD",        "Default RRSP Target Withdrawal", "dollar"),
        ("RRSP_EXTRA_DEBIT",            "RRSP Over-Draw Multiplier",      "ratio"),
        ("NONREG_EXTRA_DEBIT",          "Non-Reg Over-Draw Multiplier",   "ratio"),
        ("CAPITAL_GAIN_INCLUSION_RATE", "Capital Gains Inclusion Rate",   "percent"),
        ("OAS_CLAWBACK_THRESHOLD",      "OAS Clawback Threshold",         "dollar"),
        ("OAS_CLAWBACK_RATE",           "OAS Clawback Rate",              "percent"),
    ]),
    ("Tax Credits & Phase-Outs", [
        ("PENSION_INCOME_CREDIT_AGE",     "Pension Credit Eligible Age",    "age"),
        ("PENSION_INCOME_CREDIT_CEILING", "Pension Credit Max Income",      "dollar"),
        ("PENSION_INCOME_CREDIT_RATE",    "Pension Credit Combined Rate",   "percent"),
        ("AGE_AMOUNT_CREDIT_AGE",         "Age Amount Eligible Age",        "age"),
        ("AGE_AMOUNT_FEDERAL_BASE",       "Age Amount — Federal Base",      "dollar"),
        ("AGE_AMOUNT_FEDERAL_RATE",       "Age Amount — Federal Rate",      "percent"),
        ("AGE_AMOUNT_ONTARIO_BASE",       "Age Amount — Ontario Base",      "dollar"),
        ("AGE_AMOUNT_ONTARIO_RATE",       "Age Amount — Ontario Rate",      "percent"),
        ("AGE_AMOUNT_PHASEOUT_THRESHOLD", "Age Amount Phase-Out Threshold", "dollar"),
        ("AGE_AMOUNT_PHASEOUT_RATE",      "Age Amount Phase-Out Rate",      "percent"),
    ]),
    ("Key Milestone Ages", [
        ("LIF_MIX_AGE",       "LIF / DCPP Withdrawals Begin", "age"),
        ("PENSION_SPLIT_AGE", "Pension Income Splitting Age",  "age"),
    ]),
]


def parse_params_h(params_path):
    """Parse #define constants from params.h, returning {name: numeric_value}."""
    defines = {}
    if not os.path.exists(params_path):
        return defines
    with open(params_path, encoding="utf-8") as f:
        for line in f:
            m = re.match(r'^\s*#define\s+(\w+)\s+(.*)', line)
            if not m:
                continue
            name  = m.group(1)
            value = re.sub(r'/\*.*?\*/', '', m.group(2)).strip()
            if not value:
                continue
            try:
                # Pass already-parsed defines as locals so composite expressions
                # like (GENERICBANK_RRSP_RALPH + GENERICBROKER_RRSP_RALPH) resolve correctly.
                f_val = float(eval(value, {}, defines))
                defines[name] = int(f_val) if f_val.is_integer() else f_val
            except Exception:
                pass   # skip non-numeric / expression defines
    return defines


def build_assumptions_sheet(ws, params_path, survivor_info=None, overrides=None):
    """
    survivor_info: None for a normal run, or a dict with keys:
        'name'  — 'Ralph' or 'Sarah'
        'year'  — integer calendar year of death
        'age'   — age at death (year - yob)
    """
    COLOUR_SECTION_HDR    = "2E75B6"
    COLOUR_SURVIVOR_HDR   = "7B2D00"   # dark brick-red header
    COLOUR_SURVIVOR_ROW   = "FFF3E0"   # light amber row (matches docx event colour)

    defines = parse_params_h(params_path)
    if overrides:
        defines.update(overrides)

    # Derive work-months-in-year-1 for each spouse from retirement/current dates
    for _spouse in ("RALPH", "SARAH"):
        _ret_yr = defines.get(f"{_spouse}_RETIREMENT_YEAR")
        _cur_yr = defines.get("CURRENT_YEAR")
        _ret_mo = defines.get(f"{_spouse}_RETIREMENT_MONTH")
        _cur_mo = defines.get("CURRENT_MONTH")
        if None not in (_ret_yr, _cur_yr, _ret_mo, _cur_mo):
            _months = int(_ret_mo - _cur_mo) if _ret_yr == _cur_yr else int(_ret_mo - 1)
            defines[f"{_spouse}_WORK_MONTHS_YEAR1"] = max(_months, 0)

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 20

    row = 1

    # Title
    title = ws.cell(row=row, column=1,
                    value="Retirement Simulation \u2014 Assumptions & Parameters")
    title.font      = make_font(bold=True, size=13)
    title.alignment = Alignment(horizontal="left")
    row += 1

    generated = ws.cell(row=row, column=1,
                        value=f"Generated: {date.today().strftime('%B %d, %Y')}")
    generated.font      = make_font(size=10)
    generated.alignment = Alignment(horizontal="left")
    row += 2

    # Survivor scenario callout (shown only when --survivor was passed)
    if survivor_info:
        name      = survivor_info['name']
        surv_year = survivor_info['year']
        surv_age  = survivor_info.get('age', '')
        survivor  = 'Sarah' if name == 'Ralph' else 'Ralph'

        hdr = ws.cell(row=row, column=1, value="\u26a0  Survivor Scenario")
        hdr.font      = make_font(bold=True, color=COLOUR_HEADER_FG, size=10)
        hdr.fill      = make_fill(COLOUR_SURVIVOR_HDR)
        hdr.alignment = Alignment(horizontal="left")
        ws.cell(row=row, column=2).fill = make_fill(COLOUR_SURVIVOR_HDR)
        row += 1

        rows_data = [
            ("Deceased Spouse",   name),
            ("Year of Death",     surv_year),
            ("Age at Death",      surv_age if surv_age != '' else "—"),
            ("Surviving Spouse",  survivor),
            ("Expense Factor",    f"{int(round(survivor_info.get('expense_pct', 70)))}% of joint"),
        ]
        for label, value in rows_data:
            lbl = ws.cell(row=row, column=1, value=label)
            val = ws.cell(row=row, column=2, value=value)
            lbl.font      = make_font(size=10)
            lbl.alignment = Alignment(horizontal="left")
            val.font      = make_font(size=10, bold=(label == "Year of Death"))
            val.alignment = Alignment(horizontal="right")
            lbl.fill = val.fill = make_fill(COLOUR_SURVIVOR_ROW)
            row += 1

        row += 1   # blank spacer before first regular section

    for section_title, params in ASSUMPTION_SECTIONS:
        # Section header spanning both columns
        hdr = ws.cell(row=row, column=1, value=section_title)
        hdr.font      = make_font(bold=True, color=COLOUR_HEADER_FG, size=10)
        hdr.fill      = make_fill(COLOUR_SECTION_HDR)
        hdr.alignment = Alignment(horizontal="left")
        ws.cell(row=row, column=2).fill = make_fill(COLOUR_SECTION_HDR)
        row += 1

        for define_name, label, fmt in params:
            value = defines.get(define_name)
            if value is None:
                continue

            lbl_cell = ws.cell(row=row, column=1, value=label)
            val_cell = ws.cell(row=row, column=2, value=value)

            lbl_cell.font      = make_font(size=10)
            lbl_cell.alignment = Alignment(horizontal="left")
            val_cell.font      = make_font(size=10)
            val_cell.alignment = Alignment(horizontal="right")

            if fmt == "dollar":
                val_cell.number_format = '$#,##0'
            elif fmt == "percent":
                val_cell.number_format = '0.00%'
                # params.h stores rates as fractions (e.g. 0.05); if the value
                # is already >1 it is already a whole-number percentage — divide.
                if isinstance(value, (int, float)) and value >= 1:
                    val_cell.value = value / 100.0
            elif fmt in ("year", "age"):
                val_cell.number_format = '0'
            elif fmt == "count":
                val_cell.number_format = '#,##0'
            elif fmt == "ratio":
                # Multipliers like 1.30×, 1.10× — display with two decimal
                # places and a trailing × symbol.
                val_cell.number_format = '0.00"×"'

            if row % 2 == 0:
                lbl_cell.fill = make_fill(COLOUR_ALT_ROW)
                val_cell.fill = make_fill(COLOUR_ALT_ROW)

            row += 1

        row += 1   # blank row between sections


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def convert(csv_path, xlsx_path):
    with open(csv_path, newline="", encoding="utf-8") as f:
        # The CSV uses semicolons as delimiters
        rows = list(csv.reader(f, delimiter=";"))

    data_rows, summary_rows = split_sections(rows)

    wb = openpyxl.Workbook()

    # Parse survivor metadata written by write_simulation_to_file()
    # Column indices (0-based) for the age columns in data_rows:
    _CH_AGE_COL = 26
    _LI_AGE_COL = 36
    survivor_info = None
    for r in summary_rows:
        if r and r[0].strip() == "Survivor" and len(r) >= 3:
            name_raw = r[1].strip().lower()
            if name_raw != "none":
                surv_name = name_raw.capitalize()
                try:
                    surv_year = int(r[2].strip())
                except ValueError:
                    surv_year = 0
                # Derive age-at-death from the simulation data row for that year
                age_col   = _CH_AGE_COL if surv_name == "Ralph" else _LI_AGE_COL
                surv_age  = ""
                for dr in data_rows[1:]:    # skip header row
                    try:
                        if int(dr[0]) == surv_year and len(dr) > age_col:
                            surv_age = int(float(dr[age_col]))
                            break
                    except (ValueError, IndexError):
                        pass
                survivor_info = {
                    "name":        surv_name,
                    "year":        surv_year,
                    "age":         surv_age,
                    "expense_pct": 70,   # matches SURVIVOR_EXPENSE_FACTOR * 100
                }
            break

    # --- Detect actual moonbrook sale year from simulation data ---
    # The CSV constant MOONBROOK_SALE_YEAR is the sweep-range default, not the
    # optimised value chosen by the simulation.  Derive it the same way the
    # docx script does: the last year moonbrook rent income is non-zero, +1.
    _C_CH_RENT = 11
    _C_LI_RENT = 21
    actual_moonbrook_sale_year = None
    for dr in data_rows[1:]:
        try:
            ch_rent = float(dr[_C_CH_RENT]) if len(dr) > _C_CH_RENT else 0.0
            li_rent = float(dr[_C_LI_RENT]) if len(dr) > _C_LI_RENT else 0.0
        except (ValueError, TypeError):
            continue
        if ch_rent > 0 or li_rent > 0:
            try:
                actual_moonbrook_sale_year = int(float(dr[0])) + 1
            except (ValueError, TypeError):
                pass
    param_overrides = {}
    if actual_moonbrook_sale_year is not None:
        param_overrides["MOONBROOK_SALE_YEAR"] = actual_moonbrook_sale_year

    # --- Assumptions sheet (first tab) ---
    params_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "params.h")
    ws_assumptions = wb.active
    ws_assumptions.title = "Assumptions"
    build_assumptions_sheet(ws_assumptions, params_path, survivor_info=survivor_info,
                            overrides=param_overrides)

    # --- Simulation data sheet ---
    ws_data = wb.create_sheet("Simulation")
    build_data_sheet(ws_data, data_rows)

    # --- Summary sheet ---
    ws_summary = wb.create_sheet("Summary")
    build_summary_sheet(ws_summary, summary_rows)

    # --- Charts sheet ---
    ws_charts = wb.create_sheet("Charts")

    # Parse NW_P10/P50/P90 band rows appended after the summary block
    nw_bands = {}
    for _r in summary_rows:
        if _r and _r[0].strip() in ("NW_P10", "NW_P50", "NW_P90"):
            _tag = _r[0].strip()
            _vals = []
            for _cell in _r[1:]:
                _s = _cell.strip()
                if _s:
                    try:
                        _vals.append(float(_s))
                    except ValueError:
                        pass
            if _vals:
                nw_bands[_tag] = _vals

    build_charts_sheet(ws_charts, data_rows, wb,
                       bands=nw_bands if nw_bands else None)

    wb.save(xlsx_path)


if __name__ == "__main__":
    csv_in   = sys.argv[1] if len(sys.argv) > 1 else "retirement.csv"
    xlsx_out = sys.argv[2] if len(sys.argv) > 2 \
               else f"retirement_{date.today().strftime('%Y-%m-%d')}.xlsx"
    try:
        convert(csv_in, xlsx_out)
        print(f"Saved: {xlsx_out}")
    except FileNotFoundError as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)